#!/usr/bin/env python3

import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def serialize(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time()).isoformat()
    return value


def main():
    if len(sys.argv) < 2:
        print("Usage: read_excel_sheet_rows.py <xlsx-path>", file=sys.stderr)
        sys.exit(1)

    workbook_path = Path(sys.argv[1]).expanduser().resolve()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    result = {
        "path": str(workbook_path),
        "sheets": {},
    }

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            result["sheets"][sheet.title] = []
            continue

        headers = [
            str(value).strip() if value is not None and str(value).strip() else f"__col_{index + 1}"
            for index, value in enumerate(rows[0])
        ]

        records = []
        for row_index, row in enumerate(rows[1:], start=2):
            values = {}
            has_value = False
            for index, cell in enumerate(row):
                if index >= len(headers):
                    continue
                serialized = serialize(cell)
                if serialized not in (None, ""):
                    has_value = True
                values[headers[index]] = serialized
            if has_value:
                records.append({
                    "rowNumber": row_index,
                    "values": values,
                })

        result["sheets"][sheet.title] = records

    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
